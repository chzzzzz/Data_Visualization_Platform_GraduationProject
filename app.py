import json
import re,os
from openpyxl import Workbook
import numpy as np
from flask import Flask,request,render_template,session,redirect,jsonify
from gevent import pywsgi
from utils import query,Kshape,FCN,dtw_knn
from pymysql import *
from utils import excel_json
import settings
import datetime
import js2py
conn = connect(host='localhost', user='root',password='123456', database='chz1',port=3306)
cursor = conn.cursor()
#传入flask包
#配置SECRET_KEY
app = Flask(__name__)
app.secret_key = 'This is your session_key'

@app.route('/login',methods=['GET','POST'])
def log_in():
    if request.method == 'GET':
        return render_template('login.html')
    elif request.method == 'POST':
        request.form = dict(request.form)
        def filter_fn(item):
            return request.form['name'] in item and request.form['password'] in item

        users = query.querys('select * from user', [], 'select')
        filter_user = list(filter(filter_fn, users))

        if len(filter_user):
            session['name'] = request.form['name']
            name = session['name']
            # 寻找test_result2中最近的表名
            #找到这个表里的label分别个数
            l = query.select_result2(request.form['name'])
            if l =='no_history':
                return render_template('index_nohistory.html',name=name)
            sum = l[0]+l[1]+l[2]+l[3]
            yes = l[0]/sum*100
            no = 100-yes
            p1 = int(l[0]/sum*100)
            p2 = int(l[1]/sum*100)
            p3 = int(l[2]/sum*100)
            p4 = int(l[3]/sum*100)
            print("yes",yes)
            model,up_time= query.select_last_info(request.form['name'])
            print("type_name",type(name))
            print("model",type(model))
            if model =='dtw_knn':
                model = 'DTW+KNN'
            return render_template('index.html',yes=yes,no=no,p1=p1,p2=p2,p3=p3,p4=p4,n1=l[0],n2=l[1],n3=l[2],n4=l[3],name=name,model = model,up_time = up_time,sum=sum)
        else:
            return render_template('error.html', message="用户名与密码不匹配")

@app.route('/loginOut')
def loginOut():
    session.clear()
    return redirect('/login')

@app.route('/register',methods=['GET','POST'])
def register():
    if request.method == 'GET':
        return render_template('register.html')
    elif request.method == 'POST':
        request.form = dict(request.form)
        if request.form['password'] != request.form['passwordChecked']:
            return render_template('error.html', message='两次输入密码不同，请重新输入')
        # 检查注册email是否重复
        def filter_fn_email(item):
            return request.form['email'] in item
        def filter_fn_name(item):
            return request.form['name'] in item
        users = query.querys('select * from user',[],'select')
        filter_list_email = list(filter(filter_fn_email,users))
        filter_list_name = list(filter(filter_fn_name,users))
        if len(filter_list_email):
            return render_template('error.html', message='该邮箱已被注册')
        elif len(filter_list_name):
            return render_template('error.html', message='该用户名已被注册')
        else:
            query.querys('insert into user(email,password,name,identity) values(%s,%s,%s,%s)',[request.form['email'],request.form['password'],request.form['name'],'普通用户'])
            return redirect('/login')

@app.route('/home', methods=['GET','POST'])
def home():
    name = session.get('name')
    # 寻找test_result2中最近的表名
    # 找到这个表里的label分别个数
    l = query.select_result2(name)
    if l == 'no_history':
        return render_template('index_nohistory.html', name=name)
    sum = l[0] + l[1] + l[2] + l[3]
    yes = l[0] / sum * 100
    no = 100 - yes
    p1 = round(l[0] / sum * 100,2)
    p2 = round(l[1] / sum * 100,2)
    p3 = round(l[2] / sum * 100,2)
    p4 = round(l[3] / sum * 100,2)
    model, up_time= query.select_last_info(name)
    if model == 'dtw_knn':
        model = 'DTW+KNN'
    return render_template('index.html', yes=yes, no=no, p1=p1, p2=p2, p3=p3, p4=p4, n1=l[0], n2=l[1], n3=l[2],
                           n4=l[3], name=name, model=model, up_time=up_time,sum=sum)


@app.route('/home_2', methods=['GET','POST'])
def home_2():
    name = session.get('name')
    return render_template(
        'index_nohistory.html',
        name=name
    )
@app.route('/')
def allRequest():
    return redirect('/login')

#路由拦截，在没有登陆的时候不能访问/home等界面
#钩子函数
# 在每一次请求之前调用，这时候已经有请求了，可能在这个方法里面做请求的校验
# 如果请求的校验不成功，可以直接在此方法中进行响应，直接return之后那么就不会执行视图函数

@app.before_request
def before_request():
    pat = re.compile(r'^/static')
    if re.search(pat,request.path):
        return
    if request.path =="/login" :
        return
    if request.path =="/register" :
        return
    name = session.get('name')
    if name:
        return None
    #都不是，则访问登录页
    return redirect('/login')
#文件上传页
@app.route('/fileUpLoad',methods=['GET','POST'])
def fileUpLoad():
    name = session.get('name')
    if request.method == "GET":
        return render_template('fileUpLoad.html',name=name)
    elif request.method == "POST":
        f = request.files['file']
        path = os.path.join("./data/vis",f.filename)
        f.save(path)
        f_name = "vis_" + f.filename.split('.')[0]
        # 放到数据库里
        try:
            query.excel_input(f_name,path)
            query.record_insert(name,f_name, 'vis','','')
            times,fz_i,fz_v,nb_i,nb_v,sr_i,sr_v = query.emysql_visualize(f_name.split('.')[0])
        # with open("./j1.json", "w",encoding="utf-8") as fp:
        #     json.dump(r, fp,ensure_ascii=False)
            return render_template("fileUpload_result.html", times=times, fz_i=fz_i, fz_v=fz_v, nb_i=nb_i, nb_v=nb_v, sr_i=sr_i,sr_v=sr_v,name=name)
        except:
            return render_template('error.html',message='请检查传入文件格式')
        # json_str = json.dumps(r,ensure_ascii=False)
        # return render_template("fileUpload_result.html", fz_i=r['fz_i'],fz_v=['fz_v'],nb_i=['nb_i'],nb_v=['nb_v'],sr_i=['sr_i'],sr_v=['sr_v'],times=['times'])
        #return redirect('/fileUpLoad')
        # return jsonify({'status':'ok'})
        #return render_template("test.html", name=f.filename)

@app.route('/fileUpload_result',methods=['GET','POST'])
def result():
    name = session.get('name')
    if request.method == 'GET':
        return render_template("fileUpload_result.html",name=name)
    elif request.method == "POST":
        f = request.files['file']
        path = os.path.join("./data/vis",f.filename)
        f.save(path)
        f_name = "vis_" + f.filename.split('.')[0]
        #r = excel_json.readFile(path)
        # 放到数据库里
        query.excel_input(f_name,path)
        query.record_insert(name,f_name, 'vis','','')
        times,fz_i,fz_v,nb_i,nb_v,sr_i,sr_v = query.emysql_visualize(f.filename.split('.')[0])
        print(times)
        print(type(times))
        return render_template("fileUpload_result.html", times=times, fz_i=fz_i, fz_v=fz_v, nb_i=nb_i, nb_v=nb_v, sr_i=sr_i,sr_v=sr_v,name=name )

# 分析聚类
@app.route('/data_ana',methods=['GET','POST'])
def data_ana():
    name = session.get('name')
    return render_template("data_ana.html",name=name)

#FCN方法
@app.route('/FCN_upload',methods=['GET','POST'])
def fcn_upload():
    name = session.get('name')
    if request.method == "GET":
        return render_template('FCN_upload.html', name=name)
    else :
        try:
            f = request.files['file']
            f_name = "test_" + f.filename.split('.')[0]
            main_name = f.filename.split('.')[0]
            path = "./data/test/"+f.filename+".xlsx"
            f.save(path)
            #传入数据插入
            print('开始插入数据库')
            query.excel_input(f_name,path)
            print('插入成功数据库')
            query.record_insert(name,f_name,'test','','FCN')
            print('3')

            pre_result = FCN.use_fcn(path)
            print('4')

            # 保存本地
            def save(data, path):
                wb = Workbook()
                ws = wb.active  # 激活 worksheet
                if len(np.array(data).shape)>1:
                    [h, l] = np.array(data).shape  # h为行数，l为列数
                    for i in range(h):
                        row = []
                        for j in range(l):
                            row.append(data[i][j])
                        ws.append(row)
                else:
                    print(np.array(data).shape)
                    h =np.array(data).shape[0]
                    for i in range(h):
                        row =[]
                        row.append(data[i])
                        ws.append(row)
                wb.save(path)

            tr_name = "tr_"+f.filename.split('.')[0]
            pre_result = pre_result.tolist()
            save(pre_result, "./data/test_result/"+tr_name+".xlsx")
            # 概率结果加入数据库并记录
            query.result_input(tr_name,"./data/test_result/"+tr_name+".xlsx",4)
            query.record_insert(name,tr_name, 'test_result', f_name,'FCN')
            # result_2,状态饼图，处理pre
            pre_result = np.array(pre_result)
            index_max = np.argmax(pre_result, axis=1)
            index_max2 = np.argmax(pre_result, axis=1)
            #返回label处于0-3，处理成1-4
            index_max = (index_max+1).tolist()
            tr2_name = "tr2_"+f.filename.split('.')[0]
            tr2_path = "./data/test_result2/"+"tr2_"+f.filename.split('.')[0]+".xlsx"
            #存储result_2
            save(index_max,tr2_path)
            query.result_input(tr2_name, tr2_path, 1)
            query.record_insert(name,tr2_name, 'test_result2', f_name,'FCN')
            #可视化数组
            index_n =[0,0,0,0]
            for i in range(len(index_max)):
                if index_max[i] == 1:
                    index_n[0]=index_n[0]+1
                elif index_max[i] == 2:
                    index_n[1]=index_n[1]+1
                elif index_max[i] == 3:
                    index_n[2]=index_n[2]+1
                elif index_max[i] == 4:
                    index_n[3]=index_n[3]+1
            #平行多维图
            rawData= query.result_mysql_visualize(f_name,tr2_name)
            #barline index_max=result2,pre_result=result，window直接取5
            sql = '''select times from %s''' % (f_name)
            conn.ping(reconnect=True)
            cursor.execute(sql)
            times_t = cursor.fetchall()
            conn.commit()
            times = []
            for data in times_t:
                times.append(str(data[0]))
            # times = times.flatten().tolist()
            # 设色图proba_label
            # 字典数组  = [{'start':'','end':'','label'},]
            dic_list = []
            # 两个下标游标
            st, end = 0, 0
            while end <= len(index_max) - 1:
                if end == len(index_max) - 1:
                    dic = {
                        'start': times[st],
                        'end': times[end],
                        'label': index_max[st]
                    }
                    dic_list.append(dic)
                    break
                if index_max[st] == index_max[end]:
                    end = end + 1
                else:
                    # times[end]-1s,不然图显示不出
                    t = datetime.datetime.strptime(times[end], '%H:%M:%S')
                    delta = datetime.timedelta(seconds=1)
                    t = str(t - delta)
                    t = t.split(' ')[1]

                    dic = {
                        'start': times[st],
                        # 把end之前的都包括进去，不然上一个end和下一个start之间会有间隙
                        'end': times[end],
                        'label': index_max[st]
                    }
                    dic_list.append(dic)
                    st = end
                    end = end + 1
            pre_result=pre_result.tolist()
            pre_proba = []
            for i in range(len(pre_result)):
                pre_proba.append(pre_result[i][index_max[i]-1])
            return render_template("FCN_result.html", name=name,index_n=index_n,rawData=rawData ,pre_proba = pre_proba, times = times,main_name=main_name,dic_list=dic_list)
        except:
            return render_template('error.html',message='请检查传入文件格式')
@app.route('/FCN_result',methods=['GET','POST'])
def fcn_result():
    name = session.get('name')
    if request.method =='GET':
        return render_template("FCN_result.html", name=name)
    else :
        request.form = dict(request.form)
        print("request字典", request.form)
        re_name = request.form['re_name']
        main_name = request.form['main_name']
        print('re_name', re_name)
        print('main_name', main_name)

        # 检查re_name 是否重名
        def filter_name(item):
            return request.form['re_name'] in item

        old_names = query.querys('select input_name from result_name_map', [], 'select')
        filter_if = list(filter(filter_name, old_names))
        if len(filter_if) > 0:
            return render_template('error.html', message='该结果名称已存在，请更换')
        # 第二参数 main_name,第四参数re_name
        query.record_insert(name, main_name, 'result_name_map', re_name, 'FCN')

        js_text = '''
                window.onload = function(){
                      alert("结果保存成功");   
                      }
                '''
        js2py.eval_js(js_text)  # 执行js代码
        # 传main_name为空是因为,dtw_result.html中要main_name这个数据，这里已经经过一次re_name存储，为保证正常返回需要geimain_name一个任意值
        return redirect('/FCN_upload')


#dtw方法
@app.route('/dtw_upload',methods=['GET','POST'])
def dtw_upload():
    name = session.get('name')
    if request.method == "GET":
        return render_template("dtw_upload.html", name=name)
    else:
        try:
            f = request.files['file']
            main_name = f.filename.split('.')[0]
            f_name = "test_" + f.filename.split('.')[0]
            path = "./data/test/"+ f_name+'.xlsx'
            f.save(path)
            query.excel_input(f_name, path)
            query.record_insert(name, f_name, 'test', '','dtw_knn')
            window = 5
            m = dtw_knn.KnnDtw(n_neighbors=5, max_warping_window=window)
            pre_label,pre_proba = m.kd_predict(path,window)
            # 保存本地
            def save(data, path):
                wb = Workbook()
                ws = wb.active  # 激活 worksheet
                if len(np.array(data).shape) > 1:
                    [h, l] = np.array(data).shape  # h为行数，l为列数
                    for i in range(h):
                        row = []
                        for j in range(l):
                            row.append(data[i][j])
                        ws.append(row)
                else:
                    h = np.array(data).shape[0]
                    for i in range(h):
                        row = []
                        row.append(data[i])
                        ws.append(row)
                wb.save(path)
            tr_name = "tr_" + f.filename.split('.')[0]
            # pre_proba只有一列是label的可能性
            pre_proba = pre_proba.tolist()
            print(pre_proba)
            save(pre_proba, "./data/test_result/" + tr_name + ".xlsx")
            # 概率结果加入数据库并记录
            query.result_input(tr_name, "./data/test_result/" + tr_name + ".xlsx", 1)
            query.record_insert(name, tr_name, 'test_result', f_name,'dtw_knn')
            # result_2,状态饼图，处理pre
            pre_label = np.array(pre_label)
            index_max = pre_label.tolist()
            tr2_name = "tr2_" + f.filename.split('.')[0]
            tr2_path = "./data/test_result2/" + "tr2_" + f.filename.split('.')[0] + ".xlsx"
            # 存储result_2
            save(index_max, tr2_path)
            query.result_input(tr2_name, tr2_path, 1)
            query.record_insert(name, tr2_name, 'test_result2', f_name,'dtw_knn')
            # 可视化数组
            index_n = [0, 0, 0, 0]
            for i in range(len(index_max)):
                if index_max[i] == 1:
                    index_n[0] = index_n[0] + 1
                elif index_max[i] == 2:
                    index_n[1] = index_n[1] + 1
                elif index_max[i] == 3:
                    index_n[2] = index_n[2] + 1
                elif index_max[i] == 4:
                    index_n[3] = index_n[3] + 1
            #多维图数据
            rawData = query.result_mysql_visualize_dk(f_name, tr2_name, window)
            #prob，label数据,times,dtw只会返回每个window的标签
            sql = '''select times from(select times, ROW_NUMBER()over() rn from %s)a where a.rn %% %f=1'''%(f_name,window)
            conn.ping(reconnect=True)
            cursor.execute(sql)
            times_t = cursor.fetchall()
            conn.commit()
            times = []
            for data in times_t:
                times.append(str(data[0]))
            # times = times.flatten().tolist()
            # 设色图proba_label
            #字典数组  = [{'start':'','end':'','label'},]
            dic_list = []
            #两个下标游标
            st,end=0,0
            while end<=len(index_max)-1:
                if end ==len(index_max)-1:
                    dic = {
                        'start': times[st],
                        'end': times[end],
                        'label': index_max[st]
                    }
                    dic_list.append(dic)
                    break
                if index_max[st] ==index_max[end]:
                    end = end+1
                else:
                    #times[end]-1s,不然图显示不出
                    t = datetime.datetime.strptime(times[end], '%H:%M:%S')
                    delta = datetime.timedelta(seconds=1)
                    t = str(t-delta)
                    t = t.split(' ')[1]

                    dic ={
                        'start':times[st],
                        #把end之前的都包括进去，不然上一个end和下一个start之间会有间隙
                        'end':times[end],
                        'label':index_max[st]
                    }
                    dic_list.append(dic)
                    st = end
                    end = end+1
            print("dic_list",dic_list)
            return render_template("dtw_result.html", name=name, index_n=index_n,rawData = rawData,pre_proba = pre_proba, times = times,main_name=main_name,dic_list=dic_list)
        except:
            return render_template('error.html', message='请检查传入文件格式')

@app.route('/dtw_result',methods=['GET','POST'])
def dtw_result():
    name = session.get('name')
    if request.method=='GET':
        return render_template("dtw_result.html",name=name)
    else:
        request.form = dict(request.form)
        print("request字典",request.form)
        re_name = request.form['re_name']
        main_name = request.form['main_name']
        print('re_name',re_name)
        print('main_name',main_name)
        # 检查re_name 是否重名
        def filter_name(item):
            return request.form['re_name'] in item

        old_names = query.querys('select input_name from result_name_map', [], 'select')
        filter_if = list(filter(filter_name, old_names))
        if len(filter_if)>0:
            return render_template('error.html',message='该结果名称已存在，请更换')
        # 第二参数 main_name,第四参数re_name
        query.record_insert(name,main_name,'result_name_map',re_name,'dtw_knn')

        js_text = '''
        window.onload = function(){
              alert("结果保存成功");   
              }
        '''
        js2py.eval_js(js_text)  # 执行js代码
        #传main_name为空是因为,dtw_result.html中要main_name这个数据，这里已经经过一次re_name存储，为保证正常返回需要geimain_name一个任意值
        return redirect('/dtw_upload')

@app.route('/compare_upload',methods=['GET','POST'])
def compare_upload():
    name = session.get('name')
    if request.method == 'GET':
        return render_template("compare_upload.html",name=name)
    else:
        try:
            #dtw保存，两个模型两份test
            f = request.files['file']
            main_name_dk = f.filename.split('.')[0]+'_dk'
            f_name_dk = "test_" + main_name_dk
            path_dk = "./data/test/" + f_name_dk + '.xlsx'
            f.save(path_dk)
            query.excel_input(f_name_dk, path_dk)
            query.record_insert(name, f_name_dk, 'test', '', 'dtw_knn')
            #fcn保存
            main_name = f.filename.split('.')[0] + '_FCN'
            f_name = "test_" + main_name
            # f.save(path)
            query.excel_input(f_name, path_dk)
            query.record_insert(name, f_name,'test', '','FCN')
    #dtw预测
            window = 5
            m = dtw_knn.KnnDtw(n_neighbors=5, max_warping_window=window)
            pre_label, pre_proba_dk = m.kd_predict(path_dk, window)
            # 保存本地
            def save(data, path):
                wb = Workbook()
                ws = wb.active  # 激活 worksheet
                if len(np.array(data).shape) > 1:
                    [h, l] = np.array(data).shape  # h为行数，l为列数
                    for i in range(h):
                        row = []
                        for j in range(l):
                            row.append(data[i][j])
                        ws.append(row)
                else:
                    h = np.array(data).shape[0]
                    for i in range(h):
                        row = []
                        row.append(data[i])
                        ws.append(row)
                wb.save(path)

            tr_name_dk = "tr_" + f.filename.split('.')[0]+"_dk"
            # pre_proba只有一列是label的可能性
            pre_proba_dk = pre_proba_dk.tolist()
            save(pre_proba_dk, "./data/test_result/" + tr_name_dk + ".xlsx")
            # 概率结果加入数据库并记录
            query.result_input(tr_name_dk, "./data/test_result/" + tr_name_dk + ".xlsx", 1)
            query.record_insert(name, tr_name_dk, 'test_result', f_name, 'dtw_knn')
            # result_2,状态饼图，处理pre
            pre_label = np.array(pre_label)
            index_max = pre_label.tolist()
            tr2_name_dk = "tr2_" + f.filename.split('.')[0]+"_dk"
            tr2_path = "./data/test_result2/" + tr2_name_dk + ".xlsx"
            # 存储result_2
            save(index_max, tr2_path)
            query.result_input(tr2_name_dk, tr2_path, 1)
            query.record_insert(name, tr2_name_dk, 'test_result2', f_name, 'dtw_knn')
            # 可视化数组
            index_n_dk = [0, 0, 0, 0]
            for i in range(len(index_max)):
                if index_max[i] == 1:
                    index_n_dk[0] = index_n_dk[0] + 1
                elif index_max[i] == 2:
                    index_n_dk[1] = index_n_dk[1] + 1
                elif index_max[i] == 3:
                    index_n_dk[2] = index_n_dk[2] + 1
                elif index_max[i] == 4:
                    index_n_dk[3] = index_n_dk[3] + 1
            # prob，label数据,times,dtw只会返回每个window的标签
            sql = '''select times from(select times, ROW_NUMBER()over() rn from %s)a where a.rn %% %f=1''' % (f_name, window)
            conn.ping(reconnect=True)
            cursor.execute(sql)
            times_t = cursor.fetchall()
            conn.commit()
            times_dk = []
            for data in times_t:
                times_dk.append(str(data[0]))
            # times = times.flatten().tolist()
            # 设色图proba_label
            # 字典数组  = [{'start':'','end':'','label'},]
            dic_list_dk = []
            # 两个下标游标
            st, end = 0, 0
            while end <= len(index_max) - 1:
                if end == len(index_max) - 1:
                    dic = {
                        'start': times_dk[st],
                        'end': times_dk[end],
                        'label': index_max[st]
                    }
                    dic_list_dk.append(dic)
                    break
                if index_max[st] == index_max[end]:
                    end = end + 1
                else:
                    # times[end]-1s,不然图显示不出
                    t = datetime.datetime.strptime(times_dk[end], '%H:%M:%S')
                    delta = datetime.timedelta(seconds=1)
                    t = str(t - delta)
                    t = t.split(' ')[1]

                    dic = {
                        'start': times_dk[st],
                        # 把end之前的都包括进去，不然上一个end和下一个start之间会有间隙
                        'end': times_dk[end],
                        'label': index_max[st]
                    }
                    dic_list_dk.append(dic)
                    st = end
                    end = end + 1
            print("dic_list_dk", dic_list_dk)

        # fcn
            pre_result = FCN.use_fcn(path_dk)

            tr_name = "tr_" + f.filename.split('.')[0]+"_FCN"
            pre_result = pre_result.tolist()
            save(pre_result, "./data/test_result/" + tr_name + ".xlsx")
            # 概率结果加入数据库并记录
            query.result_input(tr_name, "./data/test_result/" + tr_name + ".xlsx", 4)
            query.record_insert(name, tr_name, 'test_result', f_name, 'FCN')
            # result_2,状态饼图，处理pre
            pre_result = np.array(pre_result)
            index_max = np.argmax(pre_result, axis=1)
            index_max2 = np.argmax(pre_result, axis=1)
            # 返回label处于0-3，处理成1-4
            index_max = (index_max + 1).tolist()
            tr2_name = "tr2_" + f.filename.split('.')[0]+"_FCN"
            tr2_path = "./data/test_result2/" + "tr2_" + f.filename.split('.')[0] + ".xlsx"
            # 存储result_2
            save(index_max, tr2_path)
            query.result_input(tr2_name, tr2_path, 1)
            query.record_insert(name, tr2_name, 'test_result2', f_name, 'FCN')
            # 可视化数组
            index_n = [0, 0, 0, 0]
            for i in range(len(index_max)):
                if index_max[i] == 1:
                    index_n[0] = index_n[0] + 1
                elif index_max[i] == 2:
                    index_n[1] = index_n[1] + 1
                elif index_max[i] == 3:
                    index_n[2] = index_n[2] + 1
                elif index_max[i] == 4:
                    index_n[3] = index_n[3] + 1
            # barline index_max=result2,pre_result=result，window直接取5
            sql = '''select times from %s''' % (f_name)
            conn.ping(reconnect=True)
            cursor.execute(sql)
            times_t = cursor.fetchall()
            conn.commit()
            times = []
            for data in times_t:
                times.append(str(data[0]))
            # times = times.flatten().tolist()
            # 设色图proba_label
            # 字典数组  = [{'start':'','end':'','label'},]
            dic_list = []
            # 两个下标游标
            st, end = 0, 0
            while end <= len(index_max) - 1:
                if end == len(index_max) - 1:
                    dic = {
                        'start': times[st],
                        'end': times[end],
                        'label': index_max[st]
                    }
                    dic_list.append(dic)
                    break
                if index_max[st] == index_max[end]:
                    end = end + 1
                else:
                    # times[end]-1s,不然图显示不出
                    t = datetime.datetime.strptime(times[end], '%H:%M:%S')
                    delta = datetime.timedelta(seconds=1)
                    t = str(t - delta)
                    t = t.split(' ')[1]

                    dic = {
                        'start': times[st],
                        # 把end之前的都包括进去，不然上一个end和下一个start之间会有间隙
                        'end': times[end],
                        'label': index_max[st]
                    }
                    dic_list.append(dic)
                    st = end
                    end = end + 1
            pre_result = pre_result.tolist()
            pre_proba = []
            for i in range(len(pre_result)):
                pre_proba.append(pre_result[i][index_max[i] - 1])

             # bar,bar_list最终传参二维数组
            bar_list = []
            labels = ['正常','全部遮挡','部分遮挡','断路']
            # pie,pie_list_dk,pie_list是最终返回参数
            pie_list_dk = []
            pie_list = []
            for i in range(len(labels)):
                row = [labels[i],index_n_dk[i]*5,index_n[i]]
                bar_list.append(row)
                row1 = {'value':index_n_dk[i],'name':labels[i]+'_DK'}
                pie_list_dk.append(row1)
                row2 = {'value':index_n[i],'name':labels[i]+'_FCN'}
                pie_list.append(row2)
            print("bar_list",bar_list)
            return render_template("compare_result.html", name=name,  main_name_dk=main_name_dk,main_name=main_name,dic_list=dic_list,pre_proba=pre_proba,
                                   times=times,  dic_list_dk=dic_list_dk,times_dk=times_dk, pre_proba_dk = pre_proba_dk,bar_list=bar_list,pie_list_dk=pie_list_dk,pie_list=pie_list)
        except:
            return render_template('error.html',message='请检查传入文件格式')

@app.route('/compare_result',methods=['GET','POST'])
def compare_result():
    name = session.get('name')
    if request.method =='GET':
        return render_template('compare_result.html',name=name)
    else:
        request.form = dict(request.form)
        print("request字典",request.form)
        re_name_dk = request.form['re_name_dk']
        re_name_fcn = request.form['re_name_fcn']
        main_name_dk = request.form['main_name_dk']
        main_name_fcn = request.form['main_name_fcn']
        print('re_name_dk',re_name_dk)
        print('main_name_fcn',main_name_fcn)
        # 检查re_name 是否重名
        def filter_name_dk(item):
            return request.form['re_name_dk'] in item

        old_names = query.querys('select input_name from result_name_map', [], 'select')
        filter_if_dk = list(filter(filter_name_dk, old_names))
        if len(filter_if_dk)>0:
            return render_template('error.html',message='该DTW+KNN结果名称已存在，请更换')
        #
        def filter_name_fcn(item):
            return request.form['re_name_fcn'] in item
        filter_if_fcn = list(filter(filter_name_fcn, old_names))
        if len(filter_if_fcn)>0:
            return render_template('error.html',message='该FCN结果名称已存在，请更换')
        # 第二参数 main_name,第四参数re_name
        query.record_insert(name,main_name_dk,'result_name_map',re_name_dk,'dtw_knn')
        query.record_insert(name,main_name_fcn,'result_name_map',re_name_fcn,'FCN')
        js_text = '''
        window.alert("结果保存成功");
        '''
        #js2py.eval_js(js_text)  # 执行js代码
        #传main_name为空是因为,dtw_result.html中要main_name这个数据，这里已经经过一次re_name存储，为保证正常返回需要geimain_name一个任意值
        return redirect('/compare_upload')

@app.route('/history',methods=['get','post'])
def history():
    name = session.get('name')
    if request.method == 'GET':
        return render_template('history.html',name=name)
    elif request.method == 'POST':
        request.form = dict(request.form)
        re_name = request.form['re_name']
        #获取结果
        sql = '''select main_name from result_name_map where input_name=\'%s\' and uname = \'%s\''''%(re_name,name)
        conn.ping(reconnect=True)
        cursor.execute(sql)
        main_name = cursor.fetchall()
        if len(main_name) == 0:
            return render_template('error.html',message='没有搜索历史记录，请确认输入！')
        main_name = main_name[0][0]
        #获取模型
        sql2 = '''select model from result_name_map where input_name=\'%s\' and uname = \'%s\''''%(re_name,name)
        conn.ping(reconnect=True)
        cursor.execute(sql2)
        model = cursor.fetchall()
        model =model[0][0]
        #判断是否有记录
        tr_name = 'tr_'+main_name
        tr2_name = 'tr2_'+main_name
        sql_1 = '''select * from %s'''%(tr_name)
        sql_2 = '''select * from %s'''%(tr2_name)
        conn.ping(reconnect=True)
        cursor.execute(sql_1)
        pre_proba = cursor.fetchall()
        conn.ping(reconnect=True)
        cursor.execute(sql_2)
        pre_label = cursor.fetchall()
        index_max = []
        for i in range(len(pre_label)):
            index_max.append(int(pre_label[i][0]))
        print(index_max)

# 饼图
        index_n = [0, 0, 0, 0]
        for i in range(len(pre_label)):
            if pre_label[i][0] == 1:
                index_n[0] = index_n[0] + 1
            elif pre_label[i][0] == 2:
                index_n[1] = index_n[1] + 1
            elif pre_label[i][0] == 3:
                index_n[2] = index_n[2] + 1
            elif pre_label[i][0] == 4:
                index_n[3] = index_n[3] + 1
# parallel
        if model =='dtw_knn':
            rawData = query.result_mysql_visualize_dk('test_'+main_name, tr2_name, 5)
        elif model =='FCN':
            rawData= query.result_mysql_visualize('test_'+main_name,tr2_name)
        f_name =  'test_'+main_name
#barline
        if model =='dtw_knn':
            # prob，label数据,times,dtw只会返回每个window的标签
            window = 5
            sql_times_dk = '''select times from(select times, ROW_NUMBER()over() rn from %s)a where a.rn %% %f=1''' % (f_name, window)
            conn.ping(reconnect=True)
            cursor.execute(sql_times_dk)
            times_t = cursor.fetchall()
            conn.commit()
            times = []
            for data in times_t:
                times.append(str(data[0]))
            # times = times.flatten().tolist()
            # 设色图proba_label
            # 字典数组  = [{'start':'','end':'','label'},]
            dic_list = []
            # 两个下标游标
            st, end = 0, 0
            while end <= len(index_max) - 1:
                if end == len(index_max) - 1:
                    dic = {
                        'start': times[st],
                        'end': times[end],
                        'label': index_max[st]
                    }
                    dic_list.append(dic)
                    break
                if index_max[st] == index_max[end]:
                    end = end + 1
                else:
                    # times[end]-1s,不然图显示不出
                    t = datetime.datetime.strptime(times[end], '%H:%M:%S')
                    delta = datetime.timedelta(seconds=1)
                    t = str(t - delta)
                    t = t.split(' ')[1]

                    dic = {
                        'start': times[st],
                        # 把end之前的都包括进去，不然上一个end和下一个start之间会有间隙
                        'end': times[end],
                        'label': index_max[st]
                    }
                    dic_list.append(dic)
                    st = end
                    end = end + 1
            pre_proba2 = []
            for i in range(len(pre_proba)):
                pre_proba2.append(pre_proba[i][0])
            print("pre_proba2", pre_proba2)
        elif model=='FCN':
            sql = '''select times from %s''' % (f_name)
            conn.ping(reconnect=True)
            cursor.execute(sql)
            times_t = cursor.fetchall()
            conn.commit()
            times = []
            for data in times_t:
                times.append(str(data[0]))
            # times = times.flatten().tolist()
            # 设色图proba_label
            # 字典数组  = [{'start':'','end':'','label'},]
            dic_list = []
            # 两个下标游标
            st, end = 0, 0
            while end <= len(index_max) - 1:
                if end == len(index_max) - 1:
                    dic = {
                        'start': times[st],
                        'end': times[end],
                        'label': index_max[st]
                    }
                    dic_list.append(dic)
                    break
                if index_max[st] == index_max[end]:
                    end = end + 1
                else:
                    # times[end]-1s,不然图显示不出
                    t = datetime.datetime.strptime(times[end], '%H:%M:%S')
                    delta = datetime.timedelta(seconds=1)
                    t = str(t - delta)
                    t = t.split(' ')[1]

                    dic = {
                        'start': times[st],
                        # 把end之前的都包括进去，不然上一个end和下一个start之间会有间隙
                        'end': times[end],
                        'label': index_max[st]
                    }
                    dic_list.append(dic)
                    st = end
                    end = end + 1
            pre_proba2 = []
            for i in range(len(pre_proba)):
                pre_proba2.append(pre_proba[i][index_max[i] - 1])
        return render_template('history_result.html', name=name, index_n=index_n,rawData=rawData,pre_proba=pre_proba2, times=times,dic_list=dic_list)

@app.route('/history_result',methods=['GET','POST'])
def history_result():
    name = session.get('name')
    if request.method == 'GET':
        return render_template("history_result.html",name=name)
# 临时测试页面
@app.route('/test',methods=['GET','POST'])
def test():
    return render_template("test.html")



# 我的
@app.route('/my_home',methods=['GET','POST'])
def my_home():
    name = session.get('name')
    if request.method=='GET':
        email, password, identity = query.user_info(name)
        return render_template("my_home.html", name=name, email=email, password=password, identity=identity)
    else:
        request.form = dict(request.form)
        if request.form['new_pass'] != request.form['new_pass_checked']:
            return render_template('error.html', message='两次输入密码不同，请重新输入')
        query.change_pass(name, request.form['new_pass'])
        email, password, identity = query.user_info(name)
        return render_template('my_home_re.html',name=name, email=email, password=password, identity=identity)

        # if status =='OK':
        #     return render_template('my_home_re.html',name=name)
        # else:
        #     return render_template('error2.html', message='密码类型有误')
@app.route('/my_home_re',methods=['GET','POST'])
def my_home_re():
    name = session.get('name')
    if request.method == 'GET':
        email, password, identity = query.user_info(name)
        return render_template("my_home_re.html", name=name, email=email, password=password, identity=identity)
    elif request == 'POST':
        request.form = dict(request.form)
        if request.form['new_pass'] != request.form['new_pass_checked']:
            return render_template('error.html', message='两次输入密码不同，请重新输入')
        query.change_pass(name, request.form['new_pass'])
        email, password, identity = query.user_info(name)
        return render_template('my_home_re.html',name=name, email=email, password=password, identity=identity)

@app.route('/error',methods=['GET','POST'])
def error():
    return render_template('error.html')
if __name__ == '__main__':
    server = pywsgi.WSGIServer(('0.0.0.0', 12345), app)
    server.serve_forever()
    app.run('192.168.0.0', 5000, debug=True)
    app.config.from_object(settings)
